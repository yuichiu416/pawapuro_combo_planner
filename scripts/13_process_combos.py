"""
Parses combos_2026-2027.xlsx and outputs combos.json and maps.json
into src/data/2026-2027/.

Sheet structure (パワプロ2026-2027 tab, rows 13-170):
  Col B : ☆ = gold special combo
  Col D : name１＆name２ (full-width or half-width ＆/&)
  Col E : location (map name)
  Col F : rewards string

maps.json structure (no max_combos — no hard limit in 2026-2027):
  {
    "ゲームセンター": { "combo_names": [[...], ...], "parent_map": null },
    "浜辺（プール）": { "combo_names": [[...], ...], "parent_map": "プール" },
    ...
  }

Usage:
    python scripts/13_process_combos.py
"""
import json
import os
import re
import sys

from openpyxl import load_workbook

current_dir = os.path.dirname(os.path.abspath(__file__))
if current_dir not in sys.path:
    sys.path.append(current_dir)

from sanitizer import sanitize_name

VERSION    = '2026-2027'
RAW_DIR    = os.path.normpath(os.path.join(current_dir, '..', 'raw_data'))
DATA_DIR   = os.path.normpath(os.path.join(current_dir, '..', 'src', 'data', VERSION))
SKILLS_FILE = os.path.normpath(os.path.join(current_dir, '..', 'src', 'data', 'skills.json'))

INPUT_FILE   = os.path.join(RAW_DIR,  f'combos_{VERSION}.xlsx')
COMBOS_OUT   = os.path.join(DATA_DIR, 'combos.json')
MAPS_OUT     = os.path.join(DATA_DIR, 'maps.json')
MAPPING_PATH = os.path.join(DATA_DIR, 'character_mapping.json')
CHARS_PATH   = os.path.join(DATA_DIR, 'characters.json')

SHEET_NAME   = f'パワプロ{VERSION}'
COMBO_START  = 13
COMBO_END    = 170

COL_GOLD    = 2   # B (1-based)
COL_COMBO   = 4   # D
COL_MAP     = 5   # E
COL_REWARD  = 6   # F

# Known submap → parent_map relationships
PARENT_MAP = {
    '海辺（プール）':  'プール',    # typo alias — user will fix to 浜辺
    '浜辺（プール）':  'プール',
    '屋台村':          'レストラン',
    '駅前':            '駅',
    '空港ロビー':      '空港',
    '球場(練習場)':    '球場(練習場)',  # standalone — no parent
}

# Location name normalisations (catches typos that survive the Excel fix)
LOCATION_ALIASES = {
    '海辺？':      '浜辺（プール）',
    '海辺（プール）': '浜辺（プール）',
    '海辺':        '浜辺（プール）',
    '場所:練習場': '練習場',
    '場所：練習場': '練習場',
}

STAT_MAP = {
    '筋力': '筋力', '敏捷': '敏捷', '技術': '技術',
    '変化球': '変化球', '精神': '精神',
}

def load_json(path):
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def split_combo(raw):
    """Split on full-width ＆ or half-width & and sanitize names."""
    parts = re.split(r'[＆&]', raw)
    return [sanitize_name(p) for p in parts if p.strip()]

def parse_rewards(reward_str, skills_db):
    result = {'skills': [], 'stats': {}, 'is_gold': False}
    if not reward_str or reward_str in ('経験点不明', '不明'):
        return result

    # Skills: anything matching "名前Lv数" or "名前lv数"
    for m in re.finditer(r'([^\d\+\-\s\t,、。！？Ll０-９0-9（）()◯○◎〇×]+)[Ll][Vv](\d+)', reward_str):
        skill_name = m.group(1).strip().rstrip('＆&')
        level      = int(m.group(2))
        if skill_name:
            is_known = skill_name in skills_db
            if not is_known:
                print(f"  🔍 Unknown skill: {skill_name!r}")
            result['skills'].append({'name': skill_name, 'level': level, 'verified': is_known})

    # Stats: 筋力30 / 敏捷+10 / 技術20 etc.
    for stat_jp, stat_key in STAT_MAP.items():
        for m in re.finditer(rf'{stat_jp}\+?(\d+)', reward_str):
            val = int(m.group(1))
            result['stats'][stat_key] = result['stats'].get(stat_key, 0) + val

    return result

def build_name_resolver(mapping_db):
    """Build a lookup that resolves partial/short names to canonical full names."""
    canonical = list(mapping_db.get('by_name', {}).keys())
    # Index: first char → list of canonical names starting with that char (fast pre-filter)
    index = {}
    for name in canonical:
        key = name[0] if name else ''
        index.setdefault(key, []).append(name)
    
    def resolve(short_name):
        if short_name in mapping_db.get('by_name', {}):
            return short_name  # already exact match
        candidates = index.get(short_name[0], []) if short_name else []
        matches = [c for c in candidates if c.startswith(short_name)]
        if len(matches) == 1:
            return matches[0]  # unambiguous prefix match
        if len(matches) > 1:
            print(f"  ⚠️  Ambiguous short name {short_name!r} → {matches} (keeping as-is)")
        return short_name  # no match or ambiguous — leave unchanged
    
    return resolve
    result = {'skills': [], 'stats': {}, 'is_gold': False}
    if not reward_str or reward_str in ('経験点不明', '不明'):
        return result

    # Skills: anything matching "名前Lv数" or "名前lv数"
    for m in re.finditer(r'([^\d\+\-\s\t,、。！？Ll０-９0-9（）()◯○◎〇×]+)[Ll][Vv](\d+)', reward_str):
        skill_name = m.group(1).strip().rstrip('＆&')
        level      = int(m.group(2))
        if skill_name:
            is_known = skill_name in skills_db
            if not is_known:
                print(f"  🔍 Unknown skill: {skill_name!r}")
            result['skills'].append({'name': skill_name, 'level': level, 'verified': is_known})

    # Stats: 筋力30 / 敏捷+10 / 技術20 etc.
    for stat_jp, stat_key in STAT_MAP.items():
        for m in re.finditer(rf'{stat_jp}\+?(\d+)', reward_str):
            val = int(m.group(1))
            result['stats'][stat_key] = result['stats'].get(stat_key, 0) + val

    return result

def parse():
    if not os.path.exists(INPUT_FILE):
        print(f"❌ File not found: {INPUT_FILE}")
        print(f"   Place the combo sheet at: {INPUT_FILE}")
        return

    skills_db  = load_json(SKILLS_FILE)
    mapping_db = load_json(MAPPING_PATH)
    chars_db   = load_json(CHARS_PATH)

    known_names = set(mapping_db.get('by_name', {}).keys())
    resolve     = build_name_resolver(mapping_db)

    wb = load_workbook(INPUT_FILE)
    if SHEET_NAME not in wb.sheetnames:
        # Fallback: use first sheet
        ws = wb.active
        print(f"⚠️  Sheet '{SHEET_NAME}' not found, using: {ws.title}")
    else:
        ws = wb[SHEET_NAME]

    combos = {}
    maps   = {}
    unmatched_names = []
    skipped = 0

    for row in ws.iter_rows(min_row=COMBO_START, max_row=COMBO_END,
                            min_col=1, max_col=COL_REWARD):
        cells = [c.value for c in row]

        def cell(col_1based):
            idx = col_1based - 1
            return str(cells[idx]).strip() if idx < len(cells) and cells[idx] else ''

        is_gold   = cell(COL_GOLD) == '☆'
        combo_raw = cell(COL_COMBO)
        map_raw   = cell(COL_MAP)
        reward_raw = cell(COL_REWARD)

        if not combo_raw or '＆' not in combo_raw and '&' not in combo_raw:
            skipped += 1
            continue

        # Normalise map name
        map_name = LOCATION_ALIASES.get(map_raw, map_raw)
        if not map_name:
            skipped += 1
            continue

        # Split and resolve partial names to canonical full names
        char_names = [resolve(n) for n in split_combo(combo_raw)]
        if len(char_names) < 2:
            print(f"  ⚠️  Could not split: {combo_raw!r}")
            skipped += 1
            continue

        # Cross-reference names
        for name in char_names:
            if known_names and name not in known_names:
                unmatched_names.append(f"{name!r} in combo {combo_raw!r}")

        combo_key = '&'.join(char_names)
        rewards   = parse_rewards(reward_raw, skills_db)
        rewards['is_gold'] = is_gold

        combos[combo_key] = {
            'characters': char_names,
            'map':        map_name,
            'rewards':    rewards,
        }

        # Build maps entry
        if map_name not in maps:
            maps[map_name] = {
                'combo_names': [],
                'parent_map':  PARENT_MAP.get(map_name, None),
            }
        maps[map_name]['combo_names'].append(char_names)

    os.makedirs(DATA_DIR, exist_ok=True)
    with open(COMBOS_OUT, 'w', encoding='utf-8') as f:
        json.dump(combos, f, ensure_ascii=False, indent=2)
    with open(MAPS_OUT, 'w', encoding='utf-8') as f:
        json.dump(maps, f, ensure_ascii=False, indent=2)

    print(f"\n[{VERSION}] ✅ {len(combos)} combos, {len(maps)} maps")
    print(f"  Maps: {list(maps.keys())}")
    print(f"  Skipped rows: {skipped}")

    if unmatched_names:
        print(f"\n  ❌ {len(unmatched_names)} unmatched names (possible typos):")
        for u in unmatched_names[:20]:
            print(f"     {u}")
        if len(unmatched_names) > 20:
            print(f"     ... and {len(unmatched_names) - 20} more")
    else:
        print(f"  ✅ All names match character_mapping.json")

if __name__ == '__main__':
    parse()
