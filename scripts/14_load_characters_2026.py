"""
Parses characters_2026-2027.xlsx and outputs characters.json
into src/data/2026-2027/.

The sheet has three character groups per row:
  A-D : Tournament characters (round / mode / school / name)
  E-F : Head-hunting characters (location / name)
  I-J : Event characters (location / name)

Cross-references names against character_mapping.json (built by
02_crawler_characters.py) to flag typos. Re-run iteratively after
fixing typos in the Excel until 0 unmatched names remain.

Usage:
    python scripts/14_load_characters_2026.py
"""
import json
import re
import os
import sys

from openpyxl import load_workbook

current_dir = os.path.dirname(os.path.abspath(__file__))
if current_dir not in sys.path:
    sys.path.append(current_dir)

VERSION      = '2026-2027'
INPUT_FILE   = os.path.normpath(os.path.join(current_dir, '..', 'raw_data', f'characters_{VERSION}.xlsx'))
DATA_DIR     = os.path.normpath(os.path.join(current_dir, '..', 'src', 'data', VERSION))
OUTPUT_PATH  = os.path.join(DATA_DIR, 'characters.json')
MAPPING_PATH = os.path.join(DATA_DIR, 'character_mapping.json')

# Col indices (0-based)
COL_ROUND    = 0   # A
COL_MODE     = 1   # B
COL_SCHOOL   = 2   # C
COL_TOUR     = 3   # D - tournament character
COL_HH_LOC  = 4   # E - headhunting location
COL_HH_NAME = 5   # F - headhunting character
COL_EV_LOC  = 8   # I - event location
COL_EV_NAME = 9   # J - event character

MODE_ALIASES = {
    'アナウンサー側': 'アナウンサー',
    'プロデューサー側': 'プロデューサー',
    '共通': '共通',
}

def clean(val):
    if not val:
        return ''
    s = str(val).strip()
    # Strip trailing notes like "※アナウンサー側限定"
    if '※' in s:
        s = s[:s.index('※')].strip()
    # Strip "No：N" / "No:N" figure numbers embedded in manager name cells
    s = re.sub(r'\s*No\s*[：:]\s*\d+.*$', '', s).strip()
    return s

def load_mapping():
    if not os.path.exists(MAPPING_PATH):
        print(f"⚠️  character_mapping.json not found at {MAPPING_PATH}")
        print("   Run 02_crawler_characters.py first for typo cross-reference.")
        return set()
    with open(MAPPING_PATH, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return set(data.get('by_name', {}).keys())

def parse():
    if not os.path.exists(INPUT_FILE):
        print(f"❌ File not found: {INPUT_FILE}")
        return

    canonical = load_mapping()
    wb = load_workbook(INPUT_FILE)
    ws = wb.active

    characters = {}
    unmatched  = []

    current_round  = ''
    current_mode   = '共通'
    current_school = ''

    # Track last seen headhunting location (persists across rows)
    last_hh_loc = ''
    last_ev_loc = ''

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        vals = [str(c.value).strip() if c.value else '' for c in row]

        def col(i): return vals[i] if i < len(vals) else ''

        a = col(COL_ROUND)
        b = col(COL_MODE)
        c = col(COL_SCHOOL)
        d = col(COL_TOUR)
        e = col(COL_HH_LOC)
        f = col(COL_HH_NAME)
        i_val = col(COL_EV_LOC)
        j_val = col(COL_EV_NAME)

        # --- Round / mode tracking ---
        if a:
            current_round = a
        if b:
            current_mode = MODE_ALIASES.get(b, b)

        # --- School header (col C non-empty, col D empty) ---
        if c and not d:
            current_school = c

        # --- Tournament character (col D) ---
        if d:
            name = clean(d)
            if name:
                entry = characters.setdefault(name, {
                    "position": "マ" if "マネ" in name else "",
                    "encounter_map": "",
                    "acquisition": "試合",
                })
                entry.update({
                    "opponent_school": current_school,
                    "round":           current_round,
                    "mode":            current_mode,
                })
                if canonical and name not in canonical:
                    unmatched.append(f"[tournament] R{row[0].row}: {name!r}")

        # --- Head-hunting character (col F) ---
        if e:
            last_hh_loc = e
        hh_name = clean(f)
        if hh_name:
            entry = characters.setdefault(hh_name, {
                "position": "マ" if "マネ" in hh_name else "",
                "opponent_school": "",
                "round": "",
                "mode": "",
            })
            entry.update({
                "encounter_map": last_hh_loc,
                "acquisition":   "ヘッドハンティング",
            })
            if canonical and hh_name not in canonical:
                unmatched.append(f"[headhunting] R{row[0].row}: {hh_name!r}")

        # --- Event character (col J) ---
        if i_val:
            last_ev_loc = i_val
        ev_name = clean(j_val)
        if ev_name:
            entry = characters.setdefault(ev_name, {
                "position": "マ" if "マネ" in ev_name else "",
                "opponent_school": "",
                "round": "",
                "mode": "",
            })
            # Event characters: set encounter_map only if not already set by headhunting
            if not entry.get("encounter_map"):
                entry["encounter_map"] = last_ev_loc
            entry.setdefault("acquisition", "イベント")
            if canonical and ev_name not in canonical:
                unmatched.append(f"[event] R{row[0].row}: {ev_name!r}")

    os.makedirs(DATA_DIR, exist_ok=True)
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(characters, f, ensure_ascii=False, indent=2)

    total = len(characters)
    print(f"\n[{VERSION}] ✅ {total} characters → {OUTPUT_PATH}")

    # Breakdown
    acq = {}
    modes = {}
    for ch in characters.values():
        a = ch.get('acquisition', '?')
        m = ch.get('mode', '') or 'N/A'
        acq[a]   = acq.get(a, 0) + 1
        modes[m] = modes.get(m, 0) + 1
    print(f"  Acquisition: {acq}")
    print(f"  Mode: {modes}")

    if total < 405:
        print(f"\n  ⚠️  Only {total}/405 characters found — sheet may be incomplete.")

    if unmatched:
        print(f"\n  ❌ {len(unmatched)} names not in character_mapping.json (possible typos):")
        for u in unmatched:
            print(f"     {u}")
    else:
        print(f"\n  ✅ All names match character_mapping.json")

if __name__ == '__main__':
    parse()