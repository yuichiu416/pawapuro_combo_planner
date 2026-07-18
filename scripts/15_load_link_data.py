"""
Parses raw_data/characters_2026-2027.xlsx (master-sheet format, pasted in from
the community walkthrough spreadsheet) into:
  - src/data/2026-2027/combos.json   (merged with existing combo entries)
  - src/data/2026-2027/links.json    (new: コツコツリンク per character)

Sheet structure is NOT a simple grid — it's driven by a state machine keyed on
cell fill color, since the sheet mixes several logically distinct sections in
one continuous column range (B:F):

  Fill color   Meaning
  ----------   -------------------------------------------------------
  FFFFF2CC     sheet title/column-label row (row 1)
  FFFFFF00     team header, 共通 route      -> TOURNAMENT section rows below
  FF00FFFF     team header, アナウンサー側 route
  FFFF9900     team header, プロデューサー側 route
  FF00FFFF/
  FFFF9900/
  FFFFFF00     also reused for the row-467 "ヘッドハンティング" section-boundary
               header, distinguished from a team header by its cell values
               being literal column labels (コンボ相手/コンボ内容) rather than
               real content.
  (no fill)    character data row, columns mean:
                 TOURNAMENT:  B=round(inherited), C=character, D=コンボ相手,
                              E=コンボ内容, F=コツコツリンク
                 HEADHUNTING: B=location(per-row, NOT inherited), C=character,
                              D=コンボ相手, E=コンボ内容, F=コツコツリンク

A run of 2+ blank rows after a recognized section ends parsing (rows past
that point are general notes/minigame tables -- out of scope, confirmed with
Roger).

Known data gaps in the source sheet as of this run (not parser bugs -- flagged
for manual review, see UNRESOLVED handling below): rows 71, 86, 87 have no
コツコツリンク value. Row 514 is a standalone special-unlock note, not a
character, and is excluded automatically (no partner, no link, no combo).

Usage:
    python scripts/15_load_link_data.py
"""
import json
import os
import re
import sys

from openpyxl import load_workbook

current_dir = os.path.dirname(os.path.abspath(__file__))
if current_dir not in sys.path:
    sys.path.append(current_dir)

from sanitizer import sanitize_name, standardize_symbols

VERSION = "2026-2027"
RAW_DIR = os.path.normpath(os.path.join(current_dir, "..", "raw_data"))
DATA_DIR = os.path.normpath(os.path.join(current_dir, "..", "src", "data", VERSION))
SKILLS_FILE = os.path.normpath(os.path.join(current_dir, "..", "src", "data", "skills.json"))

INPUT_FILE = os.path.join(RAW_DIR, f"characters_{VERSION}.xlsx")
COMBOS_OUT = os.path.join(DATA_DIR, "combos.json")
LINKS_OUT = os.path.join(DATA_DIR, "links.json")
MAPPING_PATH = os.path.join(DATA_DIR, "character_mapping.json")

ROUTE_BY_FILL = {
    "FFFFFF00": "共通",
    "FF00FFFF": "アナウンサー側",
    "FFFF9900": "プロデューサー側",
}
HEADER_FILLS = set(ROUTE_BY_FILL) | {"FFFFF2CC"}

STAT_MAP = {"筋力": "筋力", "敏捷": "敏捷", "技術": "技術", "変化球": "変化球", "精神": "精神"}


def load_json(path):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


# ---------------------------------------------------------------------------
# Reward text parsing (shared logic pattern with scripts/13_process_combos.py)
# ---------------------------------------------------------------------------

def parse_rewards(reward_str, skills_db):
    result = {"skills": [], "stats": {}}
    if not reward_str or reward_str in ("経験点不明", "不明"):
        return result

    reward_str = standardize_symbols(str(reward_str))

    for m in re.finditer(r"([^\d\+\-\s\t,、。！？Ll０-９0-9（）()]+)[Ll][Vv](\d+)", reward_str):
        skill_name = m.group(1).strip().rstrip("＆&")
        level = int(m.group(2))
        if skill_name:
            is_known = skill_name in skills_db
            result["skills"].append({"name": skill_name, "level": level, "verified": is_known})

    for stat_jp, stat_key in STAT_MAP.items():
        for m in re.finditer(rf"{stat_jp}\+?(\d+)", reward_str):
            val = int(m.group(1))
            result["stats"][stat_key] = result["stats"].get(stat_key, 0) + val

    return result


# ---------------------------------------------------------------------------
# コツコツリンク text parsing
#
# Typical shape:
#   （prereq skill(s), or-separated within a group, 、-separated between groups）、
#   condition text
#   （reward skill(s) w/ Lv）
#   ：stat rewards
#   ※ optional trailing note (spawn condition etc.)
#
# Not every entry has every part -- some entries skip the prereq/reward
# parens entirely and are just stats + a note (e.g. 京野小筆, row 7).
# The reward-skills regex and stat regex run globally over the whole string
# regardless of parens, so they still work even when the string doesn't
# follow the "normal" shape.
# ---------------------------------------------------------------------------

def parse_link_text(raw, skills_db):
    if not raw:
        return None
    text = str(raw).strip()

    # Split off trailing note. Notes are introduced by ※ and may appear
    # without a preceding space, and the sheet has at least one row with an
    # unclosed paren right before the ※ (row 7, 京野小筆) -- doesn't break
    # this split, just means note_text may retain a stray "（".
    note = None
    if "※" in text:
        text, _, note = text.partition("※")
        text = text.strip()
        note = note.strip()

    # Variant B: "（skillA、skillB）、（skillC、skillD）：stats" -- two
    # adjacent parens with no trigger verb between them at all (verified:
    # 18 rows, e.g. row 9 松崎トミオ). Neither side reliably has Lv, so this
    # can't go through the blue-prereq/gold-granted classifier below --
    # store both groups as-is and flag as a different shape.
    descriptive_match = re.match(r"^（([^（）]*)）、（([^（）]*)）[：:](.*)$", text)
    if descriptive_match:
        group1 = [s.strip() for s in descriptive_match.group(1).split("、") if s.strip()]
        group2 = [s.strip() for s in descriptive_match.group(2).split("、") if s.strip()]
        stats = parse_rewards(descriptive_match.group(3), skills_db)["stats"]
        return {
            "format": "descriptive",
            "condition": None,
            "skill_groups": [group1, group2],
            "prerequisite_skills": [],
            "granted_skills": [],
            "upgrades": [],
            "stats": stats,
            "note": note,
            "raw": str(raw).strip(),
        }

    parens = re.findall(r"（([^（）]*)）", text)
    # Paren WITH Lv/lv = the blue skill (+ level) you must already have.
    # Paren WITHOUT Lv = the gold skill you're granted when the condition is
    # met. Verified against skills.json's own type field across the whole
    # sheet: 518/539 no-Lv-paren skills are type=gold, 528/532 Lv-paren
    # skills are type=blue (the ~4% gap is exactly the variant-B rows above,
    # now handled separately).
    prereq_parens = [p for p in parens if re.search(r"[Ll][Vv]", p)]
    granted_parens = [p for p in parens if not re.search(r"[Ll][Vv]", p)]

    granted_skills = []
    for p in granted_parens:
        for part in p.split("、"):
            part = part.strip()
            if not part:
                continue
            # NOTE: \bor\b doesn't work here -- Python regex treats CJK
            # characters as \w, so there's no word boundary between "送球"
            # and "or" in "バズーカ送球orささやき戦術". Split on the literal
            # substring instead (safe: "or" doesn't appear inside any real
            # skill name in skills.json, verified separately).
            alternatives = [a.strip() for a in part.split("or") if a.strip()]
            granted_skills.append(alternatives if len(alternatives) > 1 else (alternatives[0] if alternatives else part))

    prereq_reward = parse_rewards(text, skills_db)
    prerequisite_skills = prereq_reward["skills"]

    # Pair granted[i] <-> prerequisite[i] by position when counts line up
    # (verified against 雪野楓: 内野安打◯Lv2->ロケットスタート,
    # 緩急◯Lv2->変幻自在, both position-aligned). Flat (non-"or") entries only
    # -- an "or"-group on the granted side can't be position-paired cleanly.
    upgrades = []
    flat_granted = [g for g in granted_skills if isinstance(g, str)]
    if len(flat_granted) == len(prerequisite_skills):
        for gold_name, blue in zip(flat_granted, prerequisite_skills):
            upgrades.append({
                "from": {"name": blue["name"], "level": blue["level"],
                         "type": skills_db.get(blue["name"], {}).get("type")},
                "to": {"name": gold_name, "type": skills_db.get(gold_name, {}).get("type")},
            })

    # Condition text: whatever's left after stripping all parens and the
    # leading/trailing punctuation, plus the "：stats" tail.
    condition = re.sub(r"（[^（）]*）", "", text)
    condition = condition.split("：")[0]
    condition = condition.strip("、 ,　")
    condition = condition or None

    return {
        "format": "upgrade",
        "condition": condition,
        "granted_skills": granted_skills,
        "prerequisite_skills": prerequisite_skills,
        "upgrades": upgrades,
        "stats": prereq_reward["stats"],
        "note": note,
        "raw": str(raw).strip(),
    }


# ---------------------------------------------------------------------------
# Sheet state machine
# ---------------------------------------------------------------------------

def is_header(cell):
    return cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb in HEADER_FILLS


def route_of(cell):
    if cell.fill and cell.fill.fgColor:
        return ROUTE_BY_FILL.get(cell.fill.fgColor.rgb)
    return None


def is_blank_row(vals):
    return all(v in (None, "") for v in vals)


def split_partners(raw, resolve):
    if raw is None:
        return []
    raw = str(raw).strip()
    if raw in ("", "なし"):
        return []
    parts = [resolve(sanitize_name(p.strip())) for p in raw.split("、")]
    return [p for p in parts if p]


def split_combo_blocks(raw):
    if raw is None:
        return []
    raw = str(raw).strip()
    if raw in ("", "なし"):
        return []
    return [b.strip() for b in raw.split("\n\n") if b.strip()]


def match_combo_to_partner(block, partners, character):
    """Combo block order is NOT reliably aligned to the partners column
    order (verified: most multi-partner rows are reordered) -- match by
    substring containment against the pre-➡ segment instead of position."""
    head = block.split("➡", 1)[0]
    for p in partners:
        if p in head:
            return p
    for p in partners:
        core = p[:2]
        if core and core in head:
            return p
    return None


def build_name_resolver(mapping_db):
    canonical = list(mapping_db.get("by_name", {}).keys())
    index = {}
    for name in canonical:
        key = name[0] if name else ""
        index.setdefault(key, []).append(name)

    def resolve(short_name):
        if not short_name:
            return short_name
        if short_name in mapping_db.get("by_name", {}):
            return short_name
        candidates = index.get(short_name[0], [])
        matches = [c for c in candidates if c.startswith(short_name)]
        if len(matches) == 1:
            return matches[0]
        return short_name

    return resolve


def parse_sheet(path):
    wb = load_workbook(path)
    ws = wb.active

    state = "TOURNAMENT"
    ctx = {"round": None, "school": None, "route": None}
    blank_run = 0

    tournament_rows = []
    headhunting_rows = []

    for r in range(1, ws.max_row + 1):
        b, c, d, e, f = (ws.cell(row=r, column=col).value for col in range(2, 7))
        vals = [b, c, d, e, f]

        if is_blank_row(vals):
            blank_run += 1
            if blank_run >= 2:
                state = "DONE"
            continue
        blank_run = 0

        if state == "DONE":
            continue

        b_cell, c_cell = ws.cell(row=r, column=2), ws.cell(row=r, column=3)
        if is_header(b_cell) or is_header(c_cell):
            is_section_boundary = (d == "コンボ相手" and str(e or "").startswith("コンボ内容"))
            if is_section_boundary:
                if c and "ヘッドハンティング" in str(c):
                    state = "HEADHUNTING"
                    ctx = {"round": None, "school": None, "route": None}
                continue
            if state == "TOURNAMENT":
                if b:
                    ctx["round"] = b
                if c:
                    ctx["school"] = c
                r_ = route_of(b_cell) or route_of(c_cell)
                if r_:
                    ctx["route"] = r_
            continue

        if not c:
            continue

        yield_row = {"row": r, "character_raw": str(c).strip(), "partners_raw": d,
                     "combo_raw": e, "link_raw": f}
        if state == "TOURNAMENT":
            yield_row.update(round=ctx["round"], school=ctx["school"], route=ctx["route"])
            tournament_rows.append(yield_row)
        elif state == "HEADHUNTING":
            yield_row["location"] = str(b).strip() if b else None
            headhunting_rows.append(yield_row)

    return tournament_rows, headhunting_rows


def main():
    skills_db = load_json(SKILLS_FILE)
    mapping_db = load_json(MAPPING_PATH)
    resolve = build_name_resolver(mapping_db)
    known_names = set(mapping_db.get("by_name", {}).keys())

    tournament_rows, headhunting_rows = parse_sheet(INPUT_FILE)

    existing_combos = load_json(COMBOS_OUT)
    combos = dict(existing_combos)
    links = {}
    unmatched_names = []
    manual_review = []

    def process(row, source_type):
        if not row["partners_raw"] and not row["link_raw"]:
            # No partner and no link at all -> not a real character row, just
            # stray text (e.g. row 514's standalone オリ変 unlock note).
            return

        character = resolve(sanitize_name(row["character_raw"]))
        if known_names and character not in known_names:
            unmatched_names.append(f"{character!r} (row {row['row']}, {source_type})")

        partners = split_partners(row["partners_raw"], resolve)
        combo_blocks = split_combo_blocks(row["combo_raw"])

        for block in combo_blocks:
            partner = (
                match_combo_to_partner(block, partners, character)
                if len(partners) > 1 else (partners[0] if partners else None)
            )
            if partner is None:
                manual_review.append(f"row {row['row']} ({character}): unresolved combo partner: {block[:40]!r}")
                continue
            location_match = re.match(r"^[^（]*（([^）]+)）", block)
            map_name = location_match.group(1) if location_match else None
            reward_text = block.split("➡", 1)[1] if "➡" in block else block
            rewards = parse_rewards(reward_text, skills_db)

            # Preserve whichever key ordering already exists in combos.json
            # (it's insertion-order, not sorted) to avoid double-entering the
            # same pair under both "A&B" and "B&A".
            key_fwd, key_rev = f"{character}&{partner}", f"{partner}&{character}"
            if key_rev in combos:
                combo_key = key_rev
            else:
                combo_key = key_fwd
            combos[combo_key] = {
                "characters": [character, partner],
                "map": map_name,
                "rewards": rewards,
                "source": source_type,
            }

        if row["link_raw"]:
            link = parse_link_text(row["link_raw"], skills_db)
            link["source"] = source_type
            if source_type == "tournament":
                link["round"] = row.get("round")
                link["school"] = row.get("school")
                link["route"] = row.get("route")
            else:
                link["location"] = row.get("location")
            links[character] = link
        elif not partners and not combo_blocks:
            return
        else:
            manual_review.append(f"row {row['row']} ({character}): no コツコツリンク value")

    for row in tournament_rows:
        process(row, "tournament")
    for row in headhunting_rows:
        process(row, "headhunting")

    os.makedirs(DATA_DIR, exist_ok=True)
    with open(COMBOS_OUT, "w", encoding="utf-8") as fh:
        json.dump(combos, fh, ensure_ascii=False, indent=2)
    with open(LINKS_OUT, "w", encoding="utf-8") as fh:
        json.dump(links, fh, ensure_ascii=False, indent=2)

    print(f"[{VERSION}] tournament rows: {len(tournament_rows)}, headhunting rows: {len(headhunting_rows)}")
    print(f"  combos.json: {len(combos)} total ({len(combos) - len(existing_combos)} new/updated)")
    print(f"  links.json: {len(links)} characters with コツコツリンク")

    if manual_review:
        print(f"\n  ⚠️  {len(manual_review)} rows need manual review:")
        for item in manual_review:
            print(f"     {item}")

    if unmatched_names:
        print(f"\n  🔍 {len(unmatched_names)} names not in character_mapping.json:")
        for u in unmatched_names[:20]:
            print(f"     {u}")
        if len(unmatched_names) > 20:
            print(f"     ... and {len(unmatched_names) - 20} more")


if __name__ == "__main__":
    main()
